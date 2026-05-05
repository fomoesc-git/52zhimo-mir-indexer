from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db import connect, database_path


HEADERS = [
    ("source_url", "资源链接"),
    ("title", "标题"),
    ("title_cn", "中文标题"),
    ("publisher", "出版商"),
    ("scale", "比例"),
    ("file_format", "文件格式"),
    ("paper_format", "纸张幅面"),
    ("file_size", "文件大小"),
    ("total_pages", "总页数"),
    ("download_url", "下载链接"),
    ("category", "分类"),
    ("published_at", "发布时间"),
    ("last_crawled_at", "最后采集时间"),
]


def export_dir() -> Path:
    path = database_path().parent / "exports"
    path.mkdir(parents=True, exist_ok=True)
    return path


def fetch_export_rows() -> list[dict[str, str]]:
    with connect() as conn:
        rows = conn.execute(
            """
            SELECT source_url, title, title_cn, publisher, scale, file_format,
                   paper_format, file_size, total_pages, download_url, category,
                   published_at, last_crawled_at
            FROM resources
            ORDER BY COALESCE(published_at, first_seen_at) DESC, id DESC
            """
        ).fetchall()
        return [{key: row[key] or "" for key, _ in HEADERS} for row in rows]


def export_csv() -> Path:
    path = export_dir() / "mir-modeley-resources.csv"
    rows = fetch_export_rows()
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[key for key, _ in HEADERS])
        writer.writerow({key: label for key, label in HEADERS})
        writer.writerows(rows)
    return path


def export_xlsx() -> Path:
    path = export_dir() / "mir-modeley-resources.xlsx"
    rows = fetch_export_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "资源索引"
    ws.append([label for _, label in HEADERS])

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row[key] for key, _ in HEADERS])

    widths = {
        "A": 48,
        "B": 56,
        "C": 36,
        "D": 26,
        "E": 12,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 48,
        "K": 20,
        "L": 18,
        "M": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx in (1, 10):
        col = get_column_letter(col_idx)
        for cell in ws[col][1:]:
            if cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    return path
